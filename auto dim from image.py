# -*- coding: utf-8 -*-
"""
PHẦN MỀM ĐO KÍCH THƯỚC BẢN VẼ / ẢNH & XUẤT EXCEL (ULTRA PRO)
Nâng cấp:
1. Xuất ảnh siêu nét: Phông chữ TrueType to rõ, tự động co giãn theo kích thước ảnh.
2. Thẻ ghi chú kích thước có NỀN VÀNG TRONG SUỐT 50%, chữ đen đậm chống chìm ảnh 100%.
3. Khung bao nét đỏ sắc nét, chất lượng xuất JPG/PNG 95% không vỡ hạt.
4. Live Preview 0ms Lag, Bước 2 có Ổ Khóa 🔒 sửa số đo trực tiếp.
"""

import os
import sys
import math
import subprocess

# Kích hoạt High-DPI trên Windows để chữ và nét vẽ siêu nét
try:
    import ctypes
    ctypes.windll.shcore.SetProcessDpiAwareness(2)
except:
    try:
        import ctypes
        ctypes.windll.shcore.SetProcessDpiAwareness(1)
    except:
        try:
            import ctypes
            ctypes.windll.user32.SetProcessDPIAware()
        except: pass

# Tự động cài đặt thư viện nếu thiếu
def auto_install_dependencies():
    packages = []
    try:
        from PIL import Image, ImageTk, ImageDraw, ImageFont
    except ImportError:
        packages.append("pillow")
    try:
        import openpyxl
    except ImportError:
        packages.append("openpyxl")
        
    if packages:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install"] + packages)
        except: pass

auto_install_dependencies()

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk, ImageDraw, ImageFont

HANDLE_SIZE = 8
HANDLE_RADIUS = 6
LINE_WIDTH = 2
EDGE_MARGIN = 30
PAN_SPEED = 16

class AppDoKichThuoc:
    def __init__(self, root):
        self.root = root
        self.root.title("Phần Mềm Đo Kích Thước Bản Vẽ / Ảnh & Xuất Excel (Ultra Pro)")
        self.root.geometry("1400x890")
        self.root.minsize(1050, 700)
        
        # Style giao diện phẳng hiện đại
        self.style = ttk.Style()
        try: self.style.theme_use("clam")
        except: pass
        self.style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"), background="#dfe6e9", foreground="#2d3436")
        self.style.configure("Treeview", font=("Segoe UI", 9), rowheight=26)
        
        # Dữ liệu ảnh
        self.image_path = None
        self.orig_image = None
        self.cached_photo = None
        self.cached_zoom = 1.0
        self.zoom_level = 1.0
        self.pan_offset_x = 0
        self.pan_offset_y = 0
        self.img_canvas_id = None
        
        # Mẫu chuẩn (Bước 1)
        self.calib_coords = None     # (x1, y1, x2, y2) trên ảnh gốc
        self.uniform_scale = None    # mét / pixel
        self.is_calibrated = False
        self.calib_locked = True
        self._updating_calib_ui = False
        
        # Món đo (Bước 2)
        self.meas_locked = True
        self.meas_ratio = 1.0
        self._updating_meas_ui = False
        self.selected_item_idx = None
        
        # Trạng thái tương tác chuột
        self.mode = "CALIBRATE"      # "CALIBRATE" hoặc "MEASURE"
        self.active_handle = None
        self.drag_target = None
        self.drag_start_x = 0
        self.drag_start_y = 0
        self.orig_box_before_drag = None
        
        self.temp_rect_id = None
        self.temp_text_id = None
        self.items = []
        
        self.calib_rect_id = None
        self.calib_text_id = None
        self.calib_handle_ids = {}
        
        self.init_ui()

    def init_ui(self):
        # 1. Toolbar trên cùng
        top_bar = tk.Frame(self.root, bg="#1e272e", height=50)
        top_bar.pack(side=tk.TOP, fill=tk.X)
        
        btn_open = tk.Button(top_bar, text="📂 Mở Ảnh", font=("Segoe UI", 9, "bold"),
                             bg="#00b894", fg="white", activebackground="#55efc4", activeforeground="#2d3436",
                             padx=14, pady=5, relief=tk.FLAT, cursor="hand2", command=self.open_image)
        btn_open.pack(side=tk.LEFT, padx=(12, 6), pady=8)
        
        self.btn_mode_calib = tk.Button(top_bar, text="📐 1. Mẫu Chuẩn", font=("Segoe UI", 9, "bold"),
                                        bg="#e17055", fg="white", padx=12, pady=5, relief=tk.FLAT,
                                        cursor="hand2", command=lambda: self.set_mode("CALIBRATE"))
        self.btn_mode_calib.pack(side=tk.LEFT, padx=4, pady=8)
        
        self.btn_mode_measure = tk.Button(top_bar, text="📏 2. Vẽ & Đo Đồ", font=("Segoe UI", 9, "bold"),
                                          bg="#2d3436", fg="white", padx=12, pady=5, relief=tk.FLAT,
                                          cursor="hand2", command=lambda: self.set_mode("MEASURE"))
        self.btn_mode_measure.pack(side=tk.LEFT, padx=4, pady=8)
        
        self.lbl_status = tk.Label(top_bar, text="Hãy mở một tệp ảnh để bắt đầu", font=("Segoe UI", 9),
                                   bg="#1e272e", fg="#dfe6e9")
        self.lbl_status.pack(side=tk.LEFT, padx=15, pady=8)
        
        btn_export = tk.Button(top_bar, text="📊 Xuất Excel (.xlsx)", font=("Segoe UI", 9, "bold"),
                               bg="#0984e3", fg="white", activebackground="#74b9ff",
                               padx=14, pady=5, relief=tk.FLAT, cursor="hand2", command=self.export_excel)
        btn_export.pack(side=tk.RIGHT, padx=12, pady=8)
        
        btn_export_img = tk.Button(top_bar, text="🖼️ Lưu Ảnh Có Kích Thước", font=("Segoe UI", 9, "bold"),
                                   bg="#f39c12", fg="white", padx=12, pady=5, relief=tk.FLAT,
                                   cursor="hand2", command=self.export_annotated_image)
        btn_export_img.pack(side=tk.RIGHT, padx=4, pady=8)

        # 2. Main Paned Window
        main_split = tk.PanedWindow(self.root, orient=tk.HORIZONTAL, bg="#b2bec3", sashwidth=4)
        main_split.pack(fill=tk.BOTH, expand=True)
        
        # --- CỘT TRÁI: CANVAS HIỂN THỊ ẢNH ---
        canvas_frame = tk.Frame(main_split, bg="#2d3436")
        main_split.add(canvas_frame, minsize=650)
        
        lbl_hint = tk.Label(canvas_frame, text="💡 Kéo chuột vẽ ô (Live Preview 0ms) | Đẩy chuột sát mép tự trượt view | Nhấp chọn ô để hiện 8 núm căn chỉnh",
                            font=("Segoe UI", 8), bg="#1e272e", fg="#dfe6e9", anchor="w", padx=10, pady=4)
        lbl_hint.pack(side=tk.BOTTOM, fill=tk.X)
        
        self.canvas = tk.Canvas(canvas_frame, bg="#111418", highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.canvas.bind("<ButtonPress-1>", self.on_mouse_down)
        self.canvas.bind("<B1-Motion>", self.on_mouse_move)
        self.canvas.bind("<ButtonRelease-1>", self.on_mouse_up)
        self.canvas.bind("<Motion>", self.on_mouse_hover)
        self.canvas.bind("<ButtonPress-3>", self.on_pan_start)
        self.canvas.bind("<B3-Motion>", self.on_pan_move)
        self.canvas.bind("<MouseWheel>", self.on_zoom)
        self.canvas.bind("<Button-4>", lambda e: self.zoom_image(1.15, e.x, e.y))
        self.canvas.bind("<Button-5>", lambda e: self.zoom_image(0.85, e.x, e.y))
        
        # --- CỘT PHẢI: BẢNG ĐIỀU KHIỂN & SỐ ĐO ---
        right_panel = tk.Frame(main_split, bg="#f5f6fa", width=520)
        main_split.add(right_panel, minsize=460)
        
        # ================= KHU VỰC 1: MẪU CHUẨN =================
        frame_calib = tk.LabelFrame(right_panel, text=" 📐 Bước 1: Mẫu Chuẩn & Tỉ Lệ Gốc ",
                                    font=("Segoe UI", 10, "bold"), fg="#d63031", bg="#f5f6fa", padx=12, pady=8)
        frame_calib.pack(fill=tk.X, padx=10, pady=(8, 4))
        
        calib_grid = tk.Frame(frame_calib, bg="#f5f6fa")
        calib_grid.pack(fill=tk.X)
        
        tk.Label(calib_grid, text="Dài (m):", font=("Segoe UI", 9, "bold"), bg="#f5f6fa").grid(row=0, column=0, sticky="w", pady=2)
        self.ent_calib_len = tk.Entry(calib_grid, font=("Segoe UI", 10), width=7, justify="center")
        self.ent_calib_len.insert(0, "4.00")
        self.ent_calib_len.grid(row=0, column=1, padx=4, pady=2)
        self.ent_calib_len.bind("<KeyRelease>", self.on_calib_len_change)
        
        self.btn_calib_lock = tk.Button(calib_grid, text="🔒 Khóa", font=("Segoe UI", 8, "bold"),
                                        bg="#00b894", fg="white", cursor="hand2", relief=tk.FLAT, width=6,
                                        command=self.toggle_calib_lock)
        self.btn_calib_lock.grid(row=0, column=2, padx=4, pady=2)
        
        tk.Label(calib_grid, text="Rộng/Cao (m):", font=("Segoe UI", 9, "bold"), bg="#f5f6fa").grid(row=0, column=3, sticky="w", pady=2)
        self.ent_calib_wid = tk.Entry(calib_grid, font=("Segoe UI", 10), width=7, justify="center")
        self.ent_calib_wid.insert(0, "2.00")
        self.ent_calib_wid.grid(row=0, column=4, padx=4, pady=2)
        self.ent_calib_wid.bind("<KeyRelease>", self.on_calib_wid_change)
        
        self.lbl_calib_px = tk.Label(frame_calib, text="Mẫu trên ảnh: 0 x 0 px",
                                     font=("Segoe UI", 8), fg="#636e72", bg="#f5f6fa")
        self.lbl_calib_px.pack(anchor="w", pady=(2, 3))
        
        btn_apply_calib = tk.Button(frame_calib, text="✔ Xác Nhận Mẫu Chuẩn (Tính Tỉ Lệ)", font=("Segoe UI", 9, "bold"),
                                    bg="#e17055", fg="white", cursor="hand2", relief=tk.FLAT, command=self.confirm_calibration)
        btn_apply_calib.pack(fill=tk.X, pady=(2, 2))
        
        self.lbl_calib_info = tk.Label(frame_calib, text="Chưa xác nhận tỉ lệ", font=("Segoe UI", 8, "italic"), fg="#636e72", bg="#f5f6fa")
        self.lbl_calib_info.pack(anchor="w")
        
        # ================= KHU VỰC 2: THÔNG TIN ĐỐI TƯỢNG ĐANG ĐO =================
        frame_measure = tk.LabelFrame(right_panel, text=" 📏 Bước 2: Thông tin đối tượng đo (Có Ổ Khóa 🔒) ",
                                      font=("Segoe UI", 10, "bold"), fg="#0984e3", bg="#f5f6fa", padx=12, pady=8)
        frame_measure.pack(fill=tk.X, padx=10, pady=4)
        
        meas_grid = tk.Frame(frame_measure, bg="#f5f6fa")
        meas_grid.pack(fill=tk.X)
        
        # Dòng 1: Tên món
        tk.Label(meas_grid, text="Tên món:", font=("Segoe UI", 9, "bold"), bg="#f5f6fa").grid(row=0, column=0, sticky="w", pady=3)
        self.ent_name = tk.Entry(meas_grid, font=("Segoe UI", 10), width=18)
        self.ent_name.grid(row=0, column=1, columnspan=4, sticky="we", padx=4, pady=3)
        self.ent_name.bind("<Return>", lambda e: self.apply_manual_edit_and_update())
        
        # Dòng 2: Dài (m) + Khóa + Rộng (m)
        tk.Label(meas_grid, text="Dài (m):", font=("Segoe UI", 9, "bold"), bg="#f5f6fa").grid(row=1, column=0, sticky="w", pady=3)
        self.ent_meas_len = tk.Entry(meas_grid, font=("Segoe UI", 10), width=7, justify="center")
        self.ent_meas_len.insert(0, "0.00")
        self.ent_meas_len.grid(row=1, column=1, padx=4, pady=3)
        self.ent_meas_len.bind("<KeyRelease>", self.on_meas_len_change)
        self.ent_meas_len.bind("<Return>", lambda e: self.apply_manual_edit_and_update())
        
        self.btn_meas_lock = tk.Button(meas_grid, text="🔒 Khóa", font=("Segoe UI", 8, "bold"),
                                       bg="#00b894", fg="white", cursor="hand2", relief=tk.FLAT, width=6,
                                       command=self.toggle_meas_lock)
        self.btn_meas_lock.grid(row=1, column=2, padx=4, pady=3)
        
        tk.Label(meas_grid, text="Rộng (m):", font=("Segoe UI", 9, "bold"), bg="#f5f6fa").grid(row=1, column=3, sticky="w", pady=3)
        self.ent_meas_wid = tk.Entry(meas_grid, font=("Segoe UI", 10), width=7, justify="center")
        self.ent_meas_wid.insert(0, "0.00")
        self.ent_meas_wid.grid(row=1, column=4, padx=4, pady=3)
        self.ent_meas_wid.bind("<KeyRelease>", self.on_meas_wid_change)
        self.ent_meas_wid.bind("<Return>", lambda e: self.apply_manual_edit_and_update())
        
        # Dòng 3: Chiều sâu & Preview DT/TT
        tk.Label(meas_grid, text="Chiều sâu (m):", font=("Segoe UI", 9, "bold"), bg="#f5f6fa").grid(row=2, column=0, sticky="w", pady=3)
        self.ent_depth = tk.Entry(meas_grid, font=("Segoe UI", 9), width=7, justify="center")
        self.ent_depth.insert(0, "0.00")
        self.ent_depth.grid(row=2, column=1, sticky="w", padx=4, pady=3)
        self.ent_depth.bind("<KeyRelease>", lambda e: self.update_dim_preview_label())
        self.ent_depth.bind("<Return>", lambda e: self.apply_manual_edit_and_update())
        
        self.lbl_preview_calc = tk.Label(meas_grid, text="DT = 0.00 m² | TT = 0.000 m³", font=("Segoe UI", 8, "italic"),
                                         fg="#636e72", bg="#f5f6fa")
        self.lbl_preview_calc.grid(row=2, column=2, columnspan=3, sticky="w", padx=4)
        
        btn_add = tk.Button(frame_measure, text="✔ Cập Nhật Số Đo / Tên (Enter)", font=("Segoe UI", 10, "bold"),
                            bg="#0984e3", fg="white", cursor="hand2", relief=tk.FLAT, command=self.apply_manual_edit_and_update)
        btn_add.pack(fill=tk.X, pady=(4, 2))

        # ================= KHU VỰC 3: BẢNG DANH SÁCH =================
        frame_list = tk.LabelFrame(right_panel, text=" 📋 Danh sách kích thước đã nạp ",
                                   font=("Segoe UI", 10, "bold"), fg="#2d3436", bg="#f5f6fa", padx=5, pady=5)
        frame_list.pack(fill=tk.BOTH, expand=True, padx=10, pady=4)
        
        cols = ("stt", "name", "len", "wid", "dep", "area", "vol")
        self.tree = ttk.Treeview(frame_list, columns=cols, show="headings", selectmode="browse")
        
        self.tree.heading("stt", text="STT")
        self.tree.heading("name", text="Tên đối tượng")
        self.tree.heading("len", text="Dài (m)")
        self.tree.heading("wid", text="Rộng (m)")
        self.tree.heading("dep", text="Sâu (m)")
        self.tree.heading("area", text="DT (m²)")
        self.tree.heading("vol", text="TT (m³)")
        
        self.tree.column("stt", width=35, anchor="center")
        self.tree.column("name", width=125, anchor="w")
        self.tree.column("len", width=55, anchor="center")
        self.tree.column("wid", width=55, anchor="center")
        self.tree.column("dep", width=55, anchor="center")
        self.tree.column("area", width=55, anchor="center")
        self.tree.column("vol", width=55, anchor="center")
        
        sb_y = ttk.Scrollbar(frame_list, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb_y.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tree.bind("<ButtonRelease-1>", self.on_tree_click)
        self.tree.bind("<Double-1>", self.on_edit_popup)
        
        btn_bar = tk.Frame(right_panel, bg="#f5f6fa")
        btn_bar.pack(fill=tk.X, padx=10, pady=(2, 8))
        
        tk.Button(btn_bar, text="▲ Lên", font=("Segoe UI", 9, "bold"), bg="#dfe6e9", cursor="hand2", width=7, relief=tk.FLAT, command=self.move_up).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_bar, text="▼ Xuống", font=("Segoe UI", 9, "bold"), bg="#dfe6e9", cursor="hand2", width=7, relief=tk.FLAT, command=self.move_down).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_bar, text="✏️ Sửa Tên", font=("Segoe UI", 9), bg="#dfe6e9", cursor="hand2", width=9, relief=tk.FLAT, command=self.on_edit_popup).pack(side=tk.LEFT, padx=2)
        tk.Button(btn_bar, text="🗑️ Xóa Dòng", font=("Segoe UI", 9), bg="#d63031", fg="white", cursor="hand2", width=10, relief=tk.FLAT, command=self.delete_item).pack(side=tk.RIGHT, padx=2)

    # ================== XỬ LÝ Ổ KHÓA BƯỚC 1 (MẪU CHUẨN) ==================

    def toggle_calib_lock(self):
        self.calib_locked = not self.calib_locked
        if self.calib_locked:
            self.btn_calib_lock.config(text="🔒 Khóa", bg="#00b894")
            self.sync_calib_dimensions(source="len")
        else:
            self.btn_calib_lock.config(text="🔓 Tự do", bg="#636e72")

    def on_calib_len_change(self, event=None):
        if self._updating_calib_ui or not self.calib_locked or not self.calib_coords: return
        self.sync_calib_dimensions(source="len")

    def on_calib_wid_change(self, event=None):
        if self._updating_calib_ui or not self.calib_locked or not self.calib_coords: return
        self.sync_calib_dimensions(source="wid")

    def sync_calib_dimensions(self, source="len"):
        if not self.calib_coords: return
        x1, y1, x2, y2 = self.calib_coords
        pw, ph = abs(x2 - x1), abs(y2 - y1)
        if pw <= 0 or ph <= 0: return
        ratio = max(pw, ph) / min(pw, ph)
        
        self._updating_calib_ui = True
        try:
            if source == "len":
                val = float(self.ent_calib_len.get().strip())
                if val > 0:
                    self.ent_calib_wid.delete(0, tk.END)
                    self.ent_calib_wid.insert(0, f"{val / ratio:.2f}")
            else:
                val = float(self.ent_calib_wid.get().strip())
                if val > 0:
                    self.ent_calib_len.delete(0, tk.END)
                    self.ent_calib_len.insert(0, f"{val * ratio:.2f}")
        except: pass
        self._updating_calib_ui = False

    def update_calib_pixel_info(self):
        if not self.calib_coords: return
        x1, y1, x2, y2 = self.calib_coords
        pw, ph = abs(x2 - x1), abs(y2 - y1)
        p_max, p_min = max(pw, ph), min(pw, ph)
        ratio = p_max / p_min if p_min > 0 else 1.0
        self.lbl_calib_px.config(text=f"Mẫu trên ảnh: {pw:.1f} x {ph:.1f} px | Tỉ lệ Dài/Rộng = {ratio:.2f}")
        if self.calib_locked:
            self.sync_calib_dimensions(source="len")

    # ================== XỬ LÝ Ổ KHÓA BƯỚC 2 (MÓN ĐO) ==================

    def toggle_meas_lock(self):
        self.meas_locked = not self.meas_locked
        if self.meas_locked:
            self.btn_meas_lock.config(text="🔒 Khóa", bg="#00b894")
            self.sync_meas_dimensions(source="len")
        else:
            self.btn_meas_lock.config(text="🔓 Tự do", bg="#636e72")

    def on_meas_len_change(self, event=None):
        if self._updating_meas_ui: return
        if self.meas_locked:
            self.sync_meas_dimensions(source="len")
        self.update_dim_preview_label()

    def on_meas_wid_change(self, event=None):
        if self._updating_meas_ui: return
        if self.meas_locked:
            self.sync_meas_dimensions(source="wid")
        self.update_dim_preview_label()

    def sync_meas_dimensions(self, source="len"):
        if self.meas_ratio <= 0: return
        self._updating_meas_ui = True
        try:
            if source == "len":
                val = float(self.ent_meas_len.get().strip())
                if val > 0:
                    self.ent_meas_wid.delete(0, tk.END)
                    self.ent_meas_wid.insert(0, f"{val / self.meas_ratio:.2f}")
            else:
                val = float(self.ent_meas_wid.get().strip())
                if val > 0:
                    self.ent_meas_len.delete(0, tk.END)
                    self.ent_meas_len.insert(0, f"{val * self.meas_ratio:.2f}")
        except: pass
        self._updating_meas_ui = False

    def update_dim_preview_label(self):
        try:
            d = float(self.ent_meas_len.get().strip())
            r = float(self.ent_meas_wid.get().strip())
            s = float(self.ent_depth.get().strip()) if self.ent_depth.get().strip() else 0.0
            dt = d * r
            tt = dt * s
            self.lbl_preview_calc.config(text=f"DT = {dt:.2f} m² | TT = {tt:.3f} m³")
        except: pass

    def apply_manual_edit_and_update(self):
        if self.selected_item_idx is None or not self.uniform_scale: return
        try:
            d = float(self.ent_meas_len.get().strip())
            r = float(self.ent_meas_wid.get().strip())
            s = float(self.ent_depth.get().strip()) if self.ent_depth.get().strip() else 0.0
            name = self.ent_name.get().strip()
            if d <= 0 or r <= 0: return
            
            it = self.items[self.selected_item_idx]
            if name: it["name"] = name
            it["length"] = d
            it["width"] = r
            it["depth"] = s
            it["area"] = round(d * r, 2)
            it["volume"] = round(d * r * s, 3) if s > 0 else 0.0
            
            x1, y1, x2, y2 = it["x1"], it["y1"], it["x2"], it["y2"]
            cx, cy = (x1 + x2) / 2, (y1 + y2) / 2
            pw_target = (d if abs(x2-x1) >= abs(y2-y1) else r) / self.uniform_scale
            ph_target = (r if abs(x2-x1) >= abs(y2-y1) else d) / self.uniform_scale
            
            it["x1"] = cx - pw_target / 2
            it["x2"] = cx + pw_target / 2
            it["y1"] = cy - ph_target / 2
            it["y2"] = cy + ph_target / 2
            
            self.sync_item_ui(self.selected_item_idx)
            self.update_single_tree_row(self.selected_item_idx)
        except: pass

    # ================== QUẢN LÝ ẢNH & CANVAS ==================

    def open_image(self):
        filetypes = [("Hình ảnh", "*.png *.jpg *.jpeg *.webp *.bmp"), ("Tất cả", "*.*")]
        path = filedialog.askopenfilename(title="Chọn ảnh bản vẽ / phối cảnh", filetypes=filetypes)
        if not path: return
        try:
            self.image_path = path
            self.orig_image = Image.open(path).convert("RGB")
            self.zoom_level = 1.0
            self.pan_offset_x = 0
            self.pan_offset_y = 0
            
            iw, ih = self.orig_image.size
            cx, cy = iw / 2, ih / 2
            bw, bh = iw * 0.3, ih * 0.2
            self.calib_coords = (cx - bw/2, cy - bh/2, cx + bw/2, cy + bh/2)
            
            self.items.clear()
            self.tree.delete(*self.tree.get_children())
            self.is_calibrated = False
            self.uniform_scale = None
            self.selected_item_idx = None
            self.lbl_calib_info.config(text="Chưa xác nhận tỉ lệ", fg="#636e72")
            
            self.fit_image_to_canvas()
            self.lbl_status.config(text=f"Đã nạp: {os.path.basename(path)} ({iw}x{ih}px)")
            self.set_mode("CALIBRATE")
            self.update_calib_pixel_info()
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể mở ảnh: {e}")

    def fit_image_to_canvas(self):
        if self.orig_image is None: return
        self.canvas.update()
        cw = max(self.canvas.winfo_width(), 800)
        ch = max(self.canvas.winfo_height(), 600)
        iw, ih = self.orig_image.size
        scale = min(cw / iw, ch / ih) * 0.95
        self.zoom_level = max(scale, 0.05)
        self.pan_offset_x = (cw - iw * self.zoom_level) / 2
        self.pan_offset_y = (ch - ih * self.zoom_level) / 2
        self.render_canvas(zoom_changed=True)

    def render_canvas(self, zoom_changed=False):
        if self.orig_image is None: return
        
        iw, ih = self.orig_image.size
        disp_w = max(int(iw * self.zoom_level), 1)
        disp_h = max(int(ih * self.zoom_level), 1)
        
        if zoom_changed or self.cached_photo is None:
            resized = self.orig_image.resize((disp_w, disp_h), Image.Resampling.BILINEAR)
            self.cached_photo = ImageTk.PhotoImage(resized)
            self.cached_zoom = self.zoom_level
            
            self.canvas.delete("all")
            self.img_canvas_id = self.canvas.create_image(self.pan_offset_x, self.pan_offset_y,
                                                          anchor="nw", image=self.cached_photo, tags="bg_img")
            self.recreate_all_canvas_shapes()
        else:
            self.canvas.coords(self.img_canvas_id, self.pan_offset_x, self.pan_offset_y)
            self.sync_all_shapes_coords()

    def recreate_all_canvas_shapes(self):
        for idx, it in enumerate(self.items):
            self.create_item_canvas_objects(idx, it)
        if self.calib_coords and self.mode == "CALIBRATE":
            self.create_calib_canvas_objects()

    def create_item_canvas_objects(self, idx, it):
        sx1, sy1 = self.orig_to_screen(it["x1"], it["y1"])
        sx2, sy2 = self.orig_to_screen(it["x2"], it["y2"])
        min_x, max_x = min(sx1, sx2), max(sx1, sx2)
        min_y, max_y = min(sy1, sy2), max(sy1, sy2)
        
        is_selected = (self.mode == "MEASURE" and self.selected_item_idx == idx)
        color = "#00cec9" if is_selected else "#0984e3"
        
        it['rect_id'] = self.canvas.create_rectangle(min_x, min_y, max_x, max_y, outline=color, width=LINE_WIDTH)
        lbl = f"[{it['stt']}] {it['name']} ({it['length']}x{it['width']}m)"
        it['text_id'] = self.canvas.create_text(min_x + 4, min_y + 4, anchor="nw", text=lbl,
                                               fill="#ffffff" if not is_selected else "#55efc4", font=("Segoe UI", 9, "bold"))
        
        it['handle_ids'] = {}
        if is_selected:
            handles = self.get_handle_coords(min_x, min_y, max_x, max_y)
            for h_type, (hx, hy) in handles.items():
                hid = self.canvas.create_rectangle(hx - HANDLE_SIZE/2, hy - HANDLE_SIZE/2,
                                                   hx + HANDLE_SIZE/2, hy + HANDLE_SIZE/2,
                                                   fill="#ffffff", outline="#00cec9", width=LINE_WIDTH)
                it['handle_ids'][h_type] = hid

    def create_calib_canvas_objects(self):
        cx1, cy1, cx2, cy2 = self.calib_coords
        scx1, scy1 = self.orig_to_screen(cx1, cy1)
        scx2, scy2 = self.orig_to_screen(cx2, cy2)
        min_x, max_x = min(scx1, scx2), max(scx1, scx2)
        min_y, max_y = min(scy1, scy2), max(scy1, scy2)
        
        self.calib_rect_id = self.canvas.create_rectangle(min_x, min_y, max_x, max_y, outline="#e17055", width=LINE_WIDTH, dash=(6, 3))
        self.calib_text_id = self.canvas.create_text(min_x + 6, min_y + 6, anchor="nw", text="[MẪU CHUẨN 8 NÚM]", fill="#e17055", font=("Segoe UI", 9, "bold"))
        
        self.calib_handle_ids.clear()
        handles = self.get_handle_coords(min_x, min_y, max_x, max_y)
        for h_type, (hx, hy) in handles.items():
            hid = self.canvas.create_rectangle(hx - HANDLE_SIZE/2, hy - HANDLE_SIZE/2,
                                               hx + HANDLE_SIZE/2, hy + HANDLE_SIZE/2,
                                               fill="#ffffff", outline="#e17055", width=LINE_WIDTH)
            self.calib_handle_ids[h_type] = hid

    def sync_all_shapes_coords(self):
        if self.calib_coords and self.mode == "CALIBRATE":
            self.sync_calib_ui()
        for idx in range(len(self.items)):
            self.sync_item_ui(idx)

    def sync_calib_ui(self):
        if not self.calib_coords or not self.calib_rect_id: return
        cx1, cy1, cx2, cy2 = self.calib_coords
        scx1, scy1 = self.orig_to_screen(cx1, cy1)
        scx2, scy2 = self.orig_to_screen(cx2, cy2)
        min_x, max_x = min(scx1, scx2), max(scx1, scx2)
        min_y, max_y = min(scy1, scy2), max(scy1, scy2)
        
        self.canvas.coords(self.calib_rect_id, min_x, min_y, max_x, max_y)
        self.canvas.coords(self.calib_text_id, min_x + 6, min_y + 6)
        
        handles = self.get_handle_coords(min_x, min_y, max_x, max_y)
        for h_type, (hx, hy) in handles.items():
            if h_type in self.calib_handle_ids:
                self.canvas.coords(self.calib_handle_ids[h_type],
                                   hx - HANDLE_SIZE/2, hy - HANDLE_SIZE/2,
                                   hx + HANDLE_SIZE/2, hy + HANDLE_SIZE/2)

    def sync_item_ui(self, idx):
        if not (0 <= idx < len(self.items)): return
        it = self.items[idx]
        if 'rect_id' not in it: return
        
        sx1, sy1 = self.orig_to_screen(it["x1"], it["y1"])
        sx2, sy2 = self.orig_to_screen(it["x2"], it["y2"])
        min_x, max_x = min(sx1, sx2), max(sx1, sx2)
        min_y, max_y = min(sy1, sy2), max(sy1, sy2)
        
        self.canvas.coords(it['rect_id'], min_x, min_y, max_x, max_y)
        self.canvas.coords(it['text_id'], min_x + 4, min_y + 4)
        lbl = f"[{it['stt']}] {it['name']} ({it['length']}x{it['width']}m)"
        self.canvas.itemconfig(it['text_id'], text=lbl)
        
        is_selected = (self.mode == "MEASURE" and self.selected_item_idx == idx)
        self.canvas.itemconfig(it['rect_id'], outline="#00cec9" if is_selected else "#0984e3")
        self.canvas.itemconfig(it['text_id'], fill="#55efc4" if is_selected else "#ffffff")
        
        if 'handle_ids' in it and is_selected:
            handles = self.get_handle_coords(min_x, min_y, max_x, max_y)
            for h_type, (hx, hy) in handles.items():
                if h_type in it['handle_ids']:
                    self.canvas.coords(it['handle_ids'][h_type],
                                       hx - HANDLE_SIZE/2, hy - HANDLE_SIZE/2,
                                       hx + HANDLE_SIZE/2, hy + HANDLE_SIZE/2)

    def get_handle_coords(self, min_x, min_y, max_x, max_y):
        mid_x = (min_x + max_x) / 2
        mid_y = (min_y + max_y) / 2
        return {
            'nw': (min_x, min_y), 'n':  (mid_x, min_y), 'ne': (max_x, min_y),
            'e':  (max_x, mid_y), 'se': (max_x, max_y), 's':  (mid_x, max_y),
            'sw': (min_x, max_y), 'w':  (min_x, mid_y)
        }

    def orig_to_screen(self, ox, oy):
        return ox * self.zoom_level + self.pan_offset_x, oy * self.zoom_level + self.pan_offset_y

    def screen_to_orig(self, sx, sy):
        return (sx - self.pan_offset_x) / self.zoom_level, (sy - self.pan_offset_y) / self.zoom_level

    # ================== TƯƠNG TÁC CHUỘT ==================

    def set_mode(self, mode):
        self.mode = mode
        if mode == "CALIBRATE":
            self.btn_mode_calib.config(bg="#e17055")
            self.btn_mode_measure.config(bg="#2d3436")
            self.lbl_status.config(text="Chế độ: ĐIỀU CHỈNH MẪU CHUẨN (Kéo 8 núm cam hoặc di dời khung)")
        else:
            if not self.is_calibrated:
                messagebox.showwarning("Cảnh báo", "Bạn cần bấm 'Xác Nhận Mẫu Chuẩn' ở Bước 1 trước khi đo!")
                self.set_mode("CALIBRATE")
                return
            self.btn_mode_calib.config(bg="#2d3436")
            self.btn_mode_measure.config(bg="#0984e3")
            self.lbl_status.config(text="Chế độ: ĐO ĐỐI TƯỢNG (Kéo chuột vẽ ô mới hoặc bấm chọn ô để chỉnh 8 núm)")
        self.render_canvas(zoom_changed=True)

    def get_hit_handle_for_box(self, sx, sy, x1, y1, x2, y2):
        scx1, scy1 = self.orig_to_screen(x1, y1)
        scx2, scy2 = self.orig_to_screen(x2, y2)
        min_x, max_x = min(scx1, scx2), max(scx1, scx2)
        min_y, max_y = min(scy1, scy2), max(scy1, scy2)
        
        handles = self.get_handle_coords(min_x, min_y, max_x, max_y)
        for h_type, (hx, hy) in handles.items():
            if abs(sx - hx) <= HANDLE_RADIUS + 3 and abs(sy - hy) <= HANDLE_RADIUS + 3:
                return h_type
        if min_x <= sx <= max_x and min_y <= sy <= max_y:
            return 'move'
        return None

    def find_item_at_screen(self, sx, sy):
        for idx in reversed(range(len(self.items))):
            it = self.items[idx]
            scx1, scy1 = self.orig_to_screen(it["x1"], it["y1"])
            scx2, scy2 = self.orig_to_screen(it["x2"], it["y2"])
            min_x, max_x = min(scx1, scx2), max(scx1, scx2)
            min_y, max_y = min(scy1, scy2), max(scy1, scy2)
            if min_x <= sx <= max_x and min_y <= sy <= max_y:
                return idx
        return None

    def on_mouse_hover(self, event):
        if self.orig_image is None: return
        sx, sy = event.x, event.y
        h = None
        
        if self.mode == "CALIBRATE" and self.calib_coords:
            h = self.get_hit_handle_for_box(sx, sy, *self.calib_coords)
        elif self.mode == "MEASURE" and self.selected_item_idx is not None:
            it = self.items[self.selected_item_idx]
            h = self.get_hit_handle_for_box(sx, sy, it["x1"], it["y1"], it["x2"], it["y2"])
            
        if h in ('nw', 'se'): self.canvas.config(cursor="size_nw_se")
        elif h in ('ne', 'sw'): self.canvas.config(cursor="size_ne_sw")
        elif h in ('n', 's'): self.canvas.config(cursor="size_ns")
        elif h in ('e', 'w'): self.canvas.config(cursor="size_we")
        elif h == 'move': self.canvas.config(cursor="fleur")
        else: self.canvas.config(cursor="crosshair")

    def on_mouse_down(self, event):
        if self.orig_image is None: return
        self.drag_start_x = event.x
        self.drag_start_y = event.y
        sx, sy = event.x, event.y
        
        if self.mode == "CALIBRATE":
            h = self.get_hit_handle_for_box(sx, sy, *self.calib_coords) if self.calib_coords else None
            if h:
                self.active_handle = h
                self.drag_target = 'CALIB'
                self.orig_box_before_drag = self.calib_coords
            else:
                self.active_handle = 'new_draw'
                self.drag_target = 'CALIB'
                self.temp_rect_id = self.canvas.create_rectangle(sx, sy, sx, sy, outline="#e17055", width=LINE_WIDTH, dash=(5, 3))
                self.temp_text_id = self.canvas.create_text(sx + 10, sy + 10, text="", fill="#e17055", font=("Segoe UI", 9, "bold"))
        else:
            h = None
            if self.selected_item_idx is not None:
                it = self.items[self.selected_item_idx]
                h = self.get_hit_handle_for_box(sx, sy, it["x1"], it["y1"], it["x2"], it["y2"])
                
            if h:
                self.active_handle = h
                self.drag_target = 'ITEM'
                it = self.items[self.selected_item_idx]
                self.orig_box_before_drag = (it["x1"], it["y1"], it["x2"], it["y2"])
            else:
                hit_idx = self.find_item_at_screen(sx, sy)
                if hit_idx is not None:
                    self.select_item_by_index(hit_idx)
                    it = self.items[hit_idx]
                    self.active_handle = 'move'
                    self.drag_target = 'ITEM'
                    self.orig_box_before_drag = (it["x1"], it["y1"], it["x2"], it["y2"])
                else:
                    self.selected_item_idx = None
                    self.active_handle = 'new_draw'
                    self.drag_target = 'NEW_ITEM'
                    self.temp_rect_id = self.canvas.create_rectangle(sx, sy, sx, sy, outline="#00cec9", width=LINE_WIDTH)
                    self.temp_text_id = self.canvas.create_text(sx + 10, sy + 10, text="", fill="#00cec9", font=("Segoe UI", 9, "bold"))
                    self.render_canvas(zoom_changed=True)

    def check_and_auto_pan(self, sx, sy):
        cw = max(self.canvas.winfo_width(), 100)
        ch = max(self.canvas.winfo_height(), 100)
        panned = False
        if sx < EDGE_MARGIN:
            self.pan_offset_x += PAN_SPEED
            panned = True
        elif sx > cw - EDGE_MARGIN:
            self.pan_offset_x -= PAN_SPEED
            panned = True
        if sy < EDGE_MARGIN:
            self.pan_offset_y += PAN_SPEED
            panned = True
        elif sy > ch - EDGE_MARGIN:
            self.pan_offset_y -= PAN_SPEED
            panned = True
            
        if panned:
            self.canvas.coords(self.img_canvas_id, self.pan_offset_x, self.pan_offset_y)
            self.sync_all_shapes_coords()
            return True
        return False

    def on_mouse_move(self, event):
        if self.orig_image is None: return
        sx, sy = event.x, event.y
        self.check_and_auto_pan(sx, sy)
        
        # 1. LIVE PREVIEW TỨC THÌ (0% LAG) KHI ĐANG VẼ MỚI
        if self.active_handle == 'new_draw' and self.temp_rect_id:
            min_x, max_x = min(self.drag_start_x, sx), max(self.drag_start_x, sx)
            min_y, max_y = min(self.drag_start_y, sy), max(self.drag_start_y, sy)
            self.canvas.coords(self.temp_rect_id, min_x, min_y, max_x, max_y)
            self.canvas.tag_raise(self.temp_rect_id)
            
            if self.uniform_scale and self.temp_text_id:
                ox1, oy1 = self.screen_to_orig(min_x, min_y)
                ox2, oy2 = self.screen_to_orig(max_x, max_y)
                pw, ph = abs(ox2 - ox1), abs(oy2 - oy1)
                curr_d = round(max(pw * self.uniform_scale, ph * self.uniform_scale), 2)
                curr_r = round(min(pw * self.uniform_scale, ph * self.uniform_scale), 2)
                self.canvas.coords(self.temp_text_id, max_x + 8, max_y + 8)
                self.canvas.itemconfig(self.temp_text_id, text=f"{curr_d:.2f}m x {curr_r:.2f}m")
                self.canvas.tag_raise(self.temp_text_id)
                
        elif self.active_handle and self.orig_box_before_drag:
            ox, oy = self.screen_to_orig(sx, sy)
            sox, soy = self.screen_to_orig(self.drag_start_x, self.drag_start_y)
            dx = ox - sox
            dy = oy - soy
            
            bx1, by1, bx2, by2 = self.orig_box_before_drag
            min_x, max_x = min(bx1, bx2), max(bx1, bx2)
            min_y, max_y = min(by1, by2), max(by1, by2)
            
            h = self.active_handle
            if h == 'nw':   min_x += dx; min_y += dy
            elif h == 'ne': max_x += dx; min_y += dy
            elif h == 'se': max_x += dx; max_y += dy
            elif h == 'sw': min_x += dx; max_y += dy
            elif h == 'n':  min_y += dy
            elif h == 's':  max_y += dy
            elif h == 'w':  min_x += dx
            elif h == 'e':  max_x += dx
            elif h == 'move':
                min_x += dx; max_x += dx
                min_y += dy; max_y += dy
                
            new_box = (min_x, min_y, max_x, max_y)
            
            if self.drag_target == 'CALIB':
                self.calib_coords = new_box
                self.sync_calib_ui()
                self.update_calib_pixel_info()
            elif self.drag_target == 'ITEM' and self.selected_item_idx is not None:
                self.update_item_geometry(self.selected_item_idx, new_box)
                self.sync_item_ui(self.selected_item_idx)

    def on_mouse_up(self, event):
        if self.orig_image is None: return
        sx, sy = event.x, event.y
        
        if self.active_handle == 'new_draw':
            ox1, oy1 = self.screen_to_orig(self.drag_start_x, self.drag_start_y)
            ox2, oy2 = self.screen_to_orig(sx, sy)
            x1, x2 = min(ox1, ox2), max(ox1, ox2)
            y1, y2 = min(oy1, oy2), max(oy1, oy2)
            pw, ph = x2 - x1, y2 - y1
            
            if pw > 4 and ph > 4:
                if self.drag_target == 'CALIB':
                    self.calib_coords = (x1, y1, x2, y2)
                    self.update_calib_pixel_info()
                elif self.drag_target == 'NEW_ITEM' and self.uniform_scale:
                    dai = round(max(pw * self.uniform_scale, ph * self.uniform_scale), 2)
                    rong = round(min(pw * self.uniform_scale, ph * self.uniform_scale), 2)
                    
                    stt = len(self.items) + 1
                    name = f"Món {stt:02d}"
                    depth = 0.0
                    try: depth = float(self.ent_depth.get().strip())
                    except: pass
                    
                    item = {"stt": stt, "name": name, "x1": x1, "y1": y1, "x2": x2, "y2": y2,
                            "length": dai, "width": rong, "depth": depth,
                            "area": round(dai * rong, 2), "volume": round(dai * rong * depth, 3) if depth > 0 else 0.0}
                    self.items.append(item)
                    
                    self.tree.insert("", tk.END, iid=str(stt),
                                     values=(stt, name, f"{dai:.2f}", f"{rong:.2f}", f"{depth:.2f}", f"{item['area']:.2f}", f"{item['volume']:.3f}"))
                    
                    self.select_item_by_index(len(self.items) - 1)
                    self.ent_name.focus_set()
                    self.ent_name.select_range(0, tk.END)
                    
            if self.temp_rect_id:
                self.canvas.delete(self.temp_rect_id)
                self.temp_rect_id = None
            if self.temp_text_id:
                self.canvas.delete(self.temp_text_id)
                self.temp_text_id = None
                
            self.render_canvas(zoom_changed=True)
            
        elif self.drag_target == 'ITEM' and self.selected_item_idx is not None:
            self.update_single_tree_row(self.selected_item_idx)
            
        self.active_handle = None
        self.drag_target = None

    def update_item_geometry(self, idx, box):
        x1, y1, x2, y2 = box
        min_x, max_x = min(x1, x2), max(x1, x2)
        min_y, max_y = min(y1, y2), max(y1, y2)
        pw, ph = max_x - min_x, max_y - min_y
        
        it = self.items[idx]
        it["x1"], it["y1"], it["x2"], it["y2"] = min_x, min_y, max_x, max_y
        
        if self.uniform_scale:
            dai = round(max(pw * self.uniform_scale, ph * self.uniform_scale), 2)
            rong = round(min(pw * self.uniform_scale, ph * self.uniform_scale), 2)
            it["length"] = dai
            it["width"] = rong
            it["area"] = round(dai * rong, 2)
            it["volume"] = round(dai * rong * it["depth"], 3) if it["depth"] > 0 else 0.0
            
            self.meas_ratio = dai / rong if rong > 0 else 1.0
            self._updating_meas_ui = True
            self.ent_meas_len.delete(0, tk.END)
            self.ent_meas_len.insert(0, f"{dai:.2f}")
            self.ent_meas_wid.delete(0, tk.END)
            self.ent_meas_wid.insert(0, f"{rong:.2f}")
            self._updating_meas_ui = False
            self.update_dim_preview_label()

    def update_single_tree_row(self, idx):
        if 0 <= idx < len(self.items):
            it = self.items[idx]
            self.tree.item(str(idx + 1), values=(it["stt"], it["name"], f"{it['length']:.2f}", f"{it['width']:.2f}",
                                                f"{it['depth']:.2f}", f"{it['area']:.2f}", f"{it['volume']:.3f}"))

    # --- PAN & ZOOM ---
    def on_pan_start(self, event):
        self.pan_start_x = event.x
        self.pan_start_y = event.y

    def on_pan_move(self, event):
        if self.orig_image is None: return
        self.pan_offset_x += event.x - self.pan_start_x
        self.pan_offset_y += event.y - self.pan_start_y
        self.pan_start_x = event.x
        self.pan_start_y = event.y
        self.render_canvas(zoom_changed=False)

    def on_zoom(self, event):
        if self.orig_image is None: return
        scale = 1.15 if event.delta > 0 else 0.85
        self.zoom_image(scale, event.x, event.y)

    def zoom_image(self, scale, cx, cy):
        new_zoom = self.zoom_level * scale
        if new_zoom < 0.02 or new_zoom > 30.0: return
        self.pan_offset_x = cx - (cx - self.pan_offset_x) * scale
        self.pan_offset_y = cy - (cy - self.pan_offset_y) * scale
        self.zoom_level = new_zoom
        self.render_canvas(zoom_changed=True)

    # ================== ĐIỀU KHIỂN & BẢNG SỐ ĐO ==================

    def confirm_calibration(self):
        if not self.calib_coords:
            messagebox.showwarning("Thông báo", "Chưa có khung mẫu chuẩn!")
            return
        try:
            rlen = float(self.ent_calib_len.get().strip())
            rwid = float(self.ent_calib_wid.get().strip())
            if rlen <= 0 or rwid <= 0: raise ValueError
        except:
            messagebox.showerror("Lỗi", "Kích thước Dài và Rộng mẫu phải lớn hơn 0!")
            return
            
        x1, y1, x2, y2 = self.calib_coords
        pw = abs(x2 - x1)
        ph = abs(y2 - y1)
        
        scale1 = max(rlen, rwid) / max(pw, ph)
        scale2 = min(rlen, rwid) / min(pw, ph)
        self.uniform_scale = (scale1 + scale2) / 2.0
        self.is_calibrated = True
        
        px_per_m = 1.0 / self.uniform_scale
        self.lbl_calib_info.config(text=f"✅ Đã chuẩn tỉ lệ: 1 mét ≈ {px_per_m:.1f} px", fg="#00b894")
        messagebox.showinfo("Thành công", f"Đã thiết lập tỉ lệ thành công!\n(1 mét ≈ {px_per_m:.1f} px)\nChuyển sang Bước 2 để vẽ đo các món đồ.")
        self.set_mode("MEASURE")

    def select_item_by_index(self, idx):
        if 0 <= idx < len(self.items):
            self.selected_item_idx = idx
            it = self.items[idx]
            
            self.ent_name.delete(0, tk.END)
            self.ent_name.insert(0, it["name"])
            
            self._updating_meas_ui = True
            self.ent_meas_len.delete(0, tk.END)
            self.ent_meas_len.insert(0, f"{it['length']:.2f}")
            self.ent_meas_wid.delete(0, tk.END)
            self.ent_meas_wid.insert(0, f"{it['width']:.2f}")
            self._updating_meas_ui = False
            
            self.meas_ratio = it['length'] / it['width'] if it['width'] > 0 else 1.0
            
            self.ent_depth.delete(0, tk.END)
            self.ent_depth.insert(0, f"{it['depth']:.2f}")
            self.update_dim_preview_label()
            
            self.tree.selection_set(str(idx + 1))
            self.tree.see(str(idx + 1))
            self.render_canvas(zoom_changed=True)

    def on_tree_click(self, event):
        sel = self.tree.selection()
        if not sel: return
        idx = int(sel[0]) - 1
        if 0 <= idx < len(self.items):
            self.select_item_by_index(idx)

    def delete_item(self):
        if self.selected_item_idx is None: return
        idx = self.selected_item_idx
        if 0 <= idx < len(self.items):
            del self.items[idx]
            self.selected_item_idx = None
            self.reindex_table()
            self.render_canvas(zoom_changed=True)

    def move_up(self):
        if self.selected_item_idx is None or self.selected_item_idx <= 0: return
        idx = self.selected_item_idx
        self.items[idx], self.items[idx - 1] = self.items[idx - 1], self.items[idx]
        self.selected_item_idx = idx - 1
        self.reindex_table()
        self.render_canvas(zoom_changed=True)

    def move_down(self):
        if self.selected_item_idx is None or self.selected_item_idx >= len(self.items) - 1: return
        idx = self.selected_item_idx
        self.items[idx], self.items[idx + 1] = self.items[idx + 1], self.items[idx]
        self.selected_item_idx = idx + 1
        self.reindex_table()
        self.render_canvas(zoom_changed=True)

    def reindex_table(self):
        self.tree.delete(*self.tree.get_children())
        for i, item in enumerate(self.items, 1):
            item["stt"] = i
            self.tree.insert("", tk.END, iid=str(i),
                             values=(i, item["name"], f"{item['length']:.2f}", f"{item['width']:.2f}",
                                     f"{item['depth']:.2f}", f"{item['area']:.2f}", f"{item['volume']:.3f}"))
        if self.selected_item_idx is not None and 0 <= self.selected_item_idx < len(self.items):
            self.tree.selection_set(str(self.selected_item_idx + 1))
            self.tree.see(str(self.selected_item_idx + 1))

    def on_edit_popup(self, event=None):
        if self.selected_item_idx is None: return
        item = self.items[self.selected_item_idx]
        
        top = tk.Toplevel(self.root)
        top.title("Sửa thông tin món đồ")
        top.geometry("320x170")
        top.resizable(False, False)
        top.transient(self.root)
        top.grab_set()
        
        tk.Label(top, text="Tên đối tượng:", font=("Segoe UI", 9)).pack(anchor="w", padx=15, pady=(10,2))
        e_name = tk.Entry(top, font=("Segoe UI", 10))
        e_name.insert(0, item["name"])
        e_name.pack(fill=tk.X, padx=15)
        
        tk.Label(top, text="Chiều sâu (m):", font=("Segoe UI", 9)).pack(anchor="w", padx=15, pady=(5,2))
        e_depth = tk.Entry(top, font=("Segoe UI", 10))
        e_depth.insert(0, f"{item['depth']:.2f}")
        e_depth.pack(fill=tk.X, padx=15)
        
        def save():
            new_name = e_name.get().strip()
            if new_name: item["name"] = new_name
            try:
                item["depth"] = float(e_depth.get().strip())
                item["volume"] = round(item["length"] * item["width"] * item["depth"], 3) if item["depth"] > 0 else 0.0
            except: pass
            self.update_single_tree_row(self.selected_item_idx)
            self.sync_item_ui(self.selected_item_idx)
            top.destroy()
            
        tk.Button(top, text="Lưu thay đổi", font=("Segoe UI", 9, "bold"), bg="#00b894", fg="white", command=save).pack(fill=tk.X, padx=15, pady=12)

    # ================== XUẤT EXCEL & ẢNH SIÊU NÉT (NỀN VÀNG TRONG SUỐT 50%) ==================

    def export_excel(self):
        if not self.items:
            messagebox.showwarning("Thông báo", "Danh sách đang trống!")
            return
            
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel Workbook", "*.xlsx"), ("CSV File", "*.csv")],
                                            title="Lưu bảng kích thước")
        if not path: return
        
        try:
            if path.endswith(".csv"):
                import csv
                with open(path, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f)
                    w.writerow(["STT", "Tên đối tượng", "Dài (m)", "Rộng (m)", "Sâu (m)", "Diện tích (m2)", "Thể tích (m3)"])
                    for it in self.items:
                        w.writerow([it["stt"], it["name"], it["length"], it["width"], it["depth"], it["area"], it["volume"]])
            else:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "KichThuoc"
                
                headers = ["STT", "Tên đối tượng", "Dài (m)", "Rộng (m)", "Sâu (m)", "Diện tích (m²)", "Thể tích (m³)"]
                ws.append(headers)
                
                fill = PatternFill(start_color="0984E3", end_color="0984E3", fill_type="solid")
                font_h = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                thin = Side(border_style="thin", color="D3D3D3")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                
                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=1, column=col_idx)
                    c.fill = fill
                    c.font = font_h
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = border
                
                for row_idx, it in enumerate(self.items, 2):
                    ws.append([it["stt"], it["name"], it["length"], it["width"], it["depth"], it["area"], it["volume"]])
                    for c_idx in range(1, len(headers) + 1):
                        c = ws.cell(row=row_idx, column=c_idx)
                        c.border = border
                        c.font = Font(name="Segoe UI", size=10)
                        c.alignment = Alignment(horizontal="left" if c_idx == 2 else "center", vertical="center")
                            
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                    
                wb.save(path)
                
            messagebox.showinfo("Thành công", f"Đã xuất {len(self.items)} đối tượng ra file:\n{path}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu file Excel: {e}")

    def export_annotated_image(self):
        if self.orig_image is None or not self.items:
            messagebox.showwarning("Thông báo", "Chưa có đối tượng nào để xuất ảnh!")
            return
            
        path = filedialog.asksaveasfilename(defaultextension=".jpg",
                                            filetypes=[("JPEG Image", "*.jpg"), ("PNG Image", "*.png")],
                                            title="Lưu ảnh có chú thích kích thước")
        if not path: return
        
        try:
            # Tạo overlay RGBA để vẽ trong suốt 50%
            out_img = self.orig_image.copy().convert("RGBA")
            overlay = Image.new("RGBA", out_img.size, (0, 0, 0, 0))
            draw = ImageDraw.Draw(overlay)
            
            iw, ih = out_img.size
            # Tính cỡ chữ và độ dày nét vẽ tỉ lệ theo độ phân giải ảnh
            font_size = max(int(iw / 65), 15)
            line_w = max(int(iw / 500), 2)
            
            # Load TrueType font chuẩn nét
            font = None
            for font_name in ["segoeui.ttf", "arial.ttf", "tahoma.ttf", "DejaVuSans-Bold.ttf"]:
                try:
                    font = ImageFont.truetype(font_name, font_size)
                    break
                except: pass
            if font is None:
                font = ImageFont.load_default()
                
            for it in self.items:
                x1, y1, x2, y2 = it["x1"], it["y1"], it["x2"], it["y2"]
                min_x, max_x = min(x1, x2), max(x1, x2)
                min_y, max_y = min(y1, y2), max(y1, y2)
                
                # 1. Vẽ khung chữ nhật nét đỏ đậm rõ ràng
                draw.rectangle([min_x, min_y, max_x, max_y], outline=(231, 76, 60, 255), width=line_w)
                
                # 2. Chuẩn bị nội dung nhãn
                lbl = f" [{it['stt']}] {it['name']} ({it['length']}x{it['width']}m) "
                try:
                    bbox = font.getbbox(lbl)
                    tw = bbox - bbox[0]
                    th = bbox - bbox
                except:
                    tw = len(lbl) * (font_size * 0.6)
                    th = font_size * 1.2
                    
                pad_x = int(font_size * 0.3)
                pad_y = int(font_size * 0.2)
                
                tag_x1 = min_x
                tag_y1 = min_y
                if tag_y1 + th + pad_y * 2 > ih:
                    tag_y1 = ih - th - pad_y * 2
                tag_x2 = tag_x1 + tw + pad_x * 2
                tag_y2 = tag_y1 + th + pad_y * 2
                
                # 3. Vẽ THẺ NỀN MÀU VÀNG TRONG SUỐT 50% (Alpha = 145/255)
                draw.rectangle([tag_x1, tag_y1, tag_x2, tag_y2],
                               fill=(255, 235, 59, 145), outline=(230, 126, 34, 230), width=1)
                
                # 4. Ghi chữ màu đen đậm sắc nét trên nền vàng trong suốt
                draw.text((tag_x1 + pad_x, tag_y1 + pad_y), lbl, fill=(0, 0, 0, 255), font=font)
                
            # Trộn overlay vào ảnh gốc và lưu với chất lượng 95%
            final_img = Image.alpha_composite(out_img, overlay).convert("RGB")
            final_img.save(path, quality=95)
            
            messagebox.showinfo("Thành công", f"Đã lưu ảnh sắc nét kèm nhãn nền vàng tại:\n{path}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu ảnh: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = AppDoKichThuoc(root)
    root.mainloop()